"""Excel workbook operations using openpyxl."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .validation import (
    validate_path, validate_cell_ref, validate_sheet_name, sanitize_text, validate_color,
    validate_choice,
)


def _sanitize_sheet_name(name: str) -> str:
    # Excel sheet names: max 31 chars, no [ ] : * ? / \
    name = sanitize_text(name, "sheet name")
    forbidden = '[]:*?/\\'
    cleaned = "".join("_" if c in forbidden else c for c in name)
    return cleaned[:31] or "Sheet1"


def create(
    path: str,
    sheets: dict[str, list[list[Any]]],
    *,
    header_bold: bool = True,
) -> str:
    """Create an .xlsx workbook at `path`.

    `sheets` maps sheet name -> rows (list of row lists). If empty, a single
    blank sheet named 'Sheet1' is created. The first row of each sheet is
    bolded when `header_bold` is True.
    """
    out = validate_path(path, expected_ext="xlsx", for_creation=True, operation="create")
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    if not sheets:
        sheets = {"Sheet1": []}

    for name, rows in sheets.items():
        ws = wb.create_sheet(title=_sanitize_sheet_name(name))
        for row in rows:
            ws.append([sanitize_text(str(c), "cell") if isinstance(c, str) else c for c in row])
        if header_bold and rows:
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="left")

    wb.save(str(out))
    return str(out)


def read(path: str, sheet: str | None = None) -> Any:
    """Read an .xlsx workbook.

    If `sheet` is provided, returns {'sheet': <name>, 'rows': [[...]]}.
    Otherwise returns {sheet_name: [[...]]} for every sheet.
    """
    in_ = validate_path(path, must_exist=True, expected_ext="xlsx", operation="read")
    wb = load_workbook(str(in_), data_only=True)
    if sheet is not None:
        validate_sheet_name(wb, sheet)
        ws = wb[sheet]
        return {"sheet": sheet, "rows": [list(row) for row in ws.iter_rows(values_only=True)]}
    return {
        name: [list(row) for row in wb[name].iter_rows(values_only=True)]
        for name in wb.sheetnames
    }


def append_rows(path: str, sheet: str, rows: list[list[Any]]) -> str:
    """Append rows to an existing sheet. Creates the sheet if missing."""
    in_ = validate_path(path, must_exist=True, expected_ext="xlsx", operation="append_rows")
    wb = load_workbook(str(in_))
    target = _sanitize_sheet_name(sheet)
    if target not in wb.sheetnames:
        ws = wb.create_sheet(title=target)
    else:
        ws = wb[target]
    for row in rows:
        ws.append([sanitize_text(str(c), "cell") if isinstance(c, str) else c for c in row])
    wb.save(str(in_))
    return str(in_)


def set_cell(path: str, sheet: str, cell: str, value: Any) -> str:
    """Set a single cell (A1, B2, etc.) in an existing workbook."""
    in_ = validate_path(path, must_exist=True, expected_ext="xlsx", operation="set_cell")
    cell = validate_cell_ref(cell)
    wb = load_workbook(str(in_))
    validate_sheet_name(wb, sheet)
    if isinstance(value, str):
        value = sanitize_text(value, "cell value")
    wb[sheet][cell] = value
    wb.save(str(in_))
    return str(in_)


def list_sheets(path: str) -> list[str]:
    """Return the sheet names in a workbook."""
    in_ = validate_path(path, must_exist=True, expected_ext="xlsx", operation="list_sheets")
    return load_workbook(str(in_), read_only=True).sheetnames


def delete_sheet(path: str, sheet: str) -> str:
    """Delete a sheet from an existing workbook."""
    in_ = validate_path(path, must_exist=True, expected_ext="xlsx", operation="delete_sheet")
    wb = load_workbook(str(in_))
    validate_sheet_name(wb, sheet)
    if len(wb.sheetnames) <= 1:
        raise ValueError(
            "Cannot delete the last sheet in a workbook.\n"
            "A workbook must have at least one sheet."
        )
    del wb[sheet]
    wb.save(str(in_))
    return str(in_)


def rename_sheet(path: str, old_name: str, new_name: str) -> str:
    """Rename a sheet in an existing workbook."""
    in_ = validate_path(path, must_exist=True, expected_ext="xlsx", operation="rename_sheet")
    wb = load_workbook(str(in_))
    validate_sheet_name(wb, old_name)
    if new_name in wb.sheetnames:
        raise ValueError(
            f"Sheet '{new_name}' already exists.\n"
            f"Choose a different name. Current sheets: {wb.sheetnames}"
        )
    wb[old_name].title = _sanitize_sheet_name(new_name)
    wb.save(str(in_))
    return str(in_)


def delete_rows(path: str, sheet: str, start_row: int, count: int = 1) -> str:
    """Delete rows from a sheet. ``start_row`` is 1-based."""
    in_ = validate_path(path, must_exist=True, expected_ext="xlsx", operation="delete_rows")
    wb = load_workbook(str(in_))
    validate_sheet_name(wb, sheet)
    ws = wb[sheet]
    if start_row < 1 or start_row > (ws.max_row or 1):
        raise ValueError(
            f"start_row={start_row} is out of range [1, {ws.max_row or 1}].\n"
            f"Row numbers are 1-based."
        )
    if count < 1:
        raise ValueError("count must be at least 1.")
    ws.delete_rows(start_row, count)
    wb.save(str(in_))
    return str(in_)


def merge_cells(path: str, sheet: str, range_str: str) -> str:
    """Merge cells in a range (e.g. 'A1:C1')."""
    from .validation import validate_cell_range
    in_ = validate_path(path, must_exist=True, expected_ext="xlsx", operation="merge_cells")
    range_str = validate_cell_range(range_str, "range")
    wb = load_workbook(str(in_))
    validate_sheet_name(wb, sheet)
    wb[sheet].merge_cells(range_str)
    wb.save(str(in_))
    return str(in_)


def insert_rows(path: str, sheet: str, row: int, count: int = 1) -> str:
    """Insert blank rows before ``row`` (1-based)."""
    in_ = validate_path(path, must_exist=True, expected_ext="xlsx", operation="insert_rows")
    wb = load_workbook(str(in_))
    validate_sheet_name(wb, sheet)
    ws = wb[sheet]
    if row < 1:
        raise ValueError(
            f"row={row} must be >= 1.\n"
            f"Row numbers are 1-based."
        )
    if count < 1:
        raise ValueError("count must be at least 1.")
    ws.insert_rows(row, count)
    wb.save(str(in_))
    return str(in_)


def set_column_width(path: str, sheet: str, column: str, width: float) -> str:
    """Set the width of a column (e.g. column='A', width=20)."""
    in_ = validate_path(path, must_exist=True, expected_ext="xlsx", operation="set_column_width")
    wb = load_workbook(str(in_))
    validate_sheet_name(wb, sheet)
    col = column.strip().upper()
    if not col.isalpha() or len(col) > 3:
        raise ValueError(
            f"Invalid column letter: '{column}'.\n"
            f"Use a column letter like 'A', 'B', 'AA'. Not a cell reference."
        )
    if width <= 0 or width > 255:
        raise ValueError(
            f"width={width} is out of range (0, 255].\n"
            f"Typical column widths: 8 (default), 12, 20, 30."
        )
    wb[sheet].column_dimensions[col].width = width
    wb.save(str(in_))
    return str(in_)


def freeze_panes(path: str, sheet: str, cell: str) -> str:
    """Freeze rows/columns above and to the left of ``cell``.

    For example, freeze_panes(path, sheet, 'A2') freezes the first row.
    Use 'B2' to freeze the first row and first column.
    """
    in_ = validate_path(path, must_exist=True, expected_ext="xlsx", operation="freeze_panes")
    cell = validate_cell_ref(cell)
    wb = load_workbook(str(in_))
    validate_sheet_name(wb, sheet)
    wb[sheet].freeze_panes = cell
    wb.save(str(in_))
    return str(in_)


def format_cells(
    path: str,
    sheet: str,
    cell_range: str,
    *,
    number_format: str | None = None,
    bold: bool | None = None,
    italic: bool | None = None,
    font_color: str | None = None,
    font_size: float | None = None,
    font_name: str | None = None,
    fill_color: str | None = None,
    align: str | None = None,
    valign: str | None = None,
    wrap_text: bool | None = None,
    border: str | bool | None = None,
    border_color: str = "000000",
) -> str:
    """Apply rich formatting to a cell or range (e.g. ``"A1"`` or ``"A1:C5"``).

    - ``number_format``: an Excel format code, e.g. ``"#,##0.00"``, ``"0%"``,
      ``"R#,##0.00"`` (currency), ``"yyyy-mm-dd"``.
    - Font: ``bold``/``italic``/``font_color`` (hex)/``font_size``/``font_name``.
    - ``fill_color``: cell background hex.
    - ``align`` (left/center/right) / ``valign`` (top/center/bottom) / ``wrap_text``.
    - ``border``: ``true`` or a style name (``thin``/``medium``/``thick``/``double``)
      applied to all four sides in ``border_color``.

    Existing formatting is preserved where a parameter is omitted."""
    in_ = validate_path(path, must_exist=True, expected_ext="xlsx", operation="format_cells")
    if number_format is not None and not isinstance(number_format, str):
        raise ValueError(
            f"number_format must be an Excel format-code string, got {number_format!r}.\n"
            "Examples: '#,##0.00', '0%', 'yyyy-mm-dd', 'R#,##0.00'."
        )
    align = validate_choice(align, "align", ("left", "center", "right", "justify", "general", "fill"))
    valign = validate_choice(valign, "valign", ("top", "center", "bottom", "justify"))
    border_style = None
    if border:
        border_style = validate_choice(
            border if isinstance(border, str) else "thin", "border",
            ("thin", "medium", "thick", "double", "dashed", "dotted", "hair"),
        )
    wb = load_workbook(str(in_))
    validate_sheet_name(wb, sheet)
    ws = wb[sheet]
    try:
        sel = ws[cell_range]
    except Exception as exc:
        raise ValueError(
            f"Invalid cell range '{cell_range}'. Use e.g. 'A1' or 'A1:C5'."
        ) from exc
    cells = []
    if isinstance(sel, tuple):
        for item in sel:
            if isinstance(item, tuple):
                cells.extend(item)
            else:
                cells.append(item)
    else:
        cells.append(sel)

    fc = ("FF" + validate_color(font_color)) if font_color else None
    fill = PatternFill(fill_type="solid", fgColor="FF" + validate_color(fill_color)) if fill_color else None
    side = None
    if border:
        style = border if isinstance(border, str) else "thin"
        side = Side(style=style, color="FF" + validate_color(border_color))

    for cell in cells:
        if number_format is not None:
            cell.number_format = number_format
        f = cell.font
        cell.font = Font(
            name=font_name or f.name,
            size=font_size or f.size,
            bold=bold if bold is not None else f.bold,
            italic=italic if italic is not None else f.italic,
            color=fc if fc else f.color,
        )
        if fill is not None:
            cell.fill = fill
        if align or valign or wrap_text is not None:
            a = cell.alignment
            cell.alignment = Alignment(
                horizontal=align or a.horizontal,
                vertical=valign or a.vertical,
                wrap_text=wrap_text if wrap_text is not None else a.wrap_text,
            )
        if side is not None:
            cell.border = Border(left=side, right=side, top=side, bottom=side)

    wb.save(str(in_))
    return str(in_)
